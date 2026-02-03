import openpyxl

def write_spreadsheet(filename: str, data: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "tickets"
    ws = wb["tickets"]

    i = 1
    cols = ["job_name", "cross_street", "ticket_type", "status", "id_ticket", "former_id_ticket", "release_date", "response_date", "expire_date", "permit", "days_to_expire", "cleared_ticket_date", "days_overdue"]
    for c in cols:
        ws.cell(1, i, c)
        i += 1

    row = 2
    for ticket in data:
        i = 1
        for v in ticket.values():
            ws.cell(row, i, v)
            i += 1
        row += 1

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].bestFit = True

    #print(f"Writing to {filename}.")
    wb.save(filename)



